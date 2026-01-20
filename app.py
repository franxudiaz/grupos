from flask import Flask, jsonify, request, send_file
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime

app = Flask(__name__, static_folder='public', static_url_path='')

import shutil
import glob

# Use absolute path relative to this script
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_PATH = os.path.join(BASE_DIR, 'NOVEDADES SEMANALES VEHICULOS SECCION SISTEMAS.xlsx')
BACKUP_DIR = os.path.join(BASE_DIR, 'backups')

if not os.path.exists(BACKUP_DIR):
    os.makedirs(BACKUP_DIR)

def create_daily_backup():
    """Creates a backup of the current Excel file if one doesn't exist for today."""
    if not os.path.exists(FILE_PATH):
        return
        
    today = datetime.now().strftime('%Y-%m-%d')
    backup_filename = f"backup_{today}.xlsx"
    backup_path = os.path.join(BACKUP_DIR, backup_filename)
    
    # Only create if it doesn't exist (snapshot of the start of the day)
    if not os.path.exists(backup_path):
        try:
            shutil.copy2(FILE_PATH, backup_path)
            print(f"Backup created: {backup_path}")
        except Exception as e:
            print(f"Error creating backup: {e}")


def get_sheet_type(sheet_name, df):
    # Determine type based on columns or name
    cols = [c.upper() for c in df.columns]
    
    if any('HORAS' in c for c in cols) or 'ACEITE' in cols:
        return 'grupo'
    elif 'PINES' in cols or 'RUEDAS' in cols or 'GANCHO' in cols:
        return 'remolque'
    else:
        return 'grupo'

@app.route('/')
def index():
    return app.send_static_file('index.html')

# ... (skip to restore_backup) but I can't skip in one chunk easily if they are far apart.
# I will do two separate edits to be safe.


@app.route('/api/config')
def get_config():
    if not os.path.exists(FILE_PATH):
        return jsonify({"error": "Excel file not found"}), 404
    
    xl = pd.ExcelFile(FILE_PATH)
    sheets = xl.sheet_names
    
    config = []
    for sheet in sheets:
        # Read just the header to guess type
        df = pd.read_excel(FILE_PATH, sheet_name=sheet, nrows=0)
        
        # Clean column names (strip whitespace)
        df.columns = df.columns.str.strip()
        
        type_ = get_sheet_type(sheet, df)
        config.append({"name": sheet, "type": type_})
        
    return jsonify({"sheets": config})

@app.route('/api/submit', methods=['POST'])
def submit_review():
    # 1. Ensure we have a backup before modifying
    create_daily_backup()

    data = request.json
    sheet_name = data.get('sheetName')
    form_data = data.get('formData')
    
    if not sheet_name or not form_data:
        return jsonify({"error": "Missing data"}), 400
        
    try:
        wb = load_workbook(FILE_PATH)
        if sheet_name not in wb.sheetnames:
             return jsonify({"error": "Sheet not found"}), 404
             
        ws = wb[sheet_name]
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        headers = [str(h).strip() if h else '' for h in headers]
        
        # Prepare row data
        new_row = []
        for header in headers:
            # Map header to form data key (normalized)
            # The prompt says: "GRUPO": null (in prev json), "FECHA", "HORAS", etc.
            # We need to match the Excel headers exactly.
            
            key = header.upper()
            val = None
            
            # Special logic for FECHA if not provided or auto
            if 'FECHA' in key:
                val = datetime.now().strftime('%d/%m/%Y')
            elif key in form_data:
                val = form_data[key]
            else:
                 # Try to find case-insensitive match
                 for k, v in form_data.items():
                     if k.upper() == key:
                         val = v
                         break
            
            # Use empty string if None to avoid literal "None" in excel
            if val is None:
                val = ""
            elif isinstance(val, str):
                val = val.upper()
                
            new_row.append(val)
            
        ws.append(new_row)
        
        # Apply Styling to the new row
        from openpyxl.styles import Alignment, Border, Side
        
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
                             
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # The new row index is ws.max_row
        row_idx = ws.max_row
        for col_idx in range(1, len(new_row) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center_alignment
            cell.border = thin_border

        wb.save(FILE_PATH)
        
        return jsonify({"success": True})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/delete_last', methods=['POST'])
def delete_last_row():
    # 1. Ensure we have a backup before modifying
    create_daily_backup()

    data = request.json
    sheet_name = data.get('sheetName')
    
    if not sheet_name:
        return jsonify({"error": "Missing sheet name"}), 400
        
    try:
        wb = load_workbook(FILE_PATH)
        if sheet_name not in wb.sheetnames:
             return jsonify({"error": "Sheet not found"}), 404
             
        ws = wb[sheet_name]
        
        if ws.max_row > 1: # Assuming header is row 1, don't delete header
            ws.delete_rows(ws.max_row)
            wb.save(FILE_PATH)
            return jsonify({"success": True, "message": "Última fila eliminada"})
        else:
            return jsonify({"error": "No hay filas para borrar (solo cabecera)"}), 400
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/download')
def download_file():
    if not os.path.exists(FILE_PATH):
        return jsonify({"error": "File not found"}), 404
    return send_file(FILE_PATH, as_attachment=True)

# --- BACKUP ENDPOINTS ---

@app.route('/api/backups', methods=['GET'])
def list_backups():
    files = glob.glob(os.path.join(BACKUP_DIR, '*.xlsx'))
    # Sort by creation time (newest first)
    files.sort(key=os.path.getmtime, reverse=True)
    
    backup_list = []
    for f in files:
        name = os.path.basename(f)
        timestamp = os.path.getmtime(f)
        date_str = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
        backup_list.append({"name": name, "date": date_str})
        
    return jsonify({"backups": backup_list})

@app.route('/api/restore', methods=['POST'])
def restore_backup():
    data = request.json
    backup_name = data.get('backupName')
    
    if not backup_name:
        return jsonify({"error": "Missing backup name"}), 400
        
    backup_path = os.path.join(BACKUP_DIR, backup_name)
    if not os.path.exists(backup_path):
        return jsonify({"error": "Backup not found"}), 404
        
    try:
        # Restore (User requested NO safety backup made here)
        shutil.copy2(backup_path, FILE_PATH)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
